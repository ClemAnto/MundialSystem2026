# -*- coding: utf-8 -*-
"""
Genera Mondiali2026_Scommesse.xlsx (da caricare su Google Drive -> Foglio Google).
Fogli:
  - Scommesse : tutte le selezioni delle 9 bollette + valutazione live
  - Classifiche: 12 gironi x 4 squadre (composizione ufficiale), classifica automatica, icone obblighi
  - Riepilogo : stato (vincente/perdente/definitivo) per bolletta
  - Config    : modalita' SIM + timestamp aggiornamento (usata dall'Apps Script SofaScore)
  - Calcoli   : helper (nascosto)
  - Istruzioni: guida setup live + simulazione
I risultati (V/N/P/GF/GS, celle gialle) li scrive l'Apps Script da SofaScore;
in modalita' SIM lo script non sovrascrive e puoi modificarli a mano.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- DATI BOLLETTE
F3 = [("F", "Olanda", "Giappone", 2.25), ("F", "Olanda", "Svezia", 2.75), ("F", "Giappone", "Svezia", 5.00)]
D3 = [("D", "Stati Uniti", "Turchia", 2.25), ("D", "Stati Uniti", "Paraguay", 3.50), ("D", "Turchia", "Paraguay", 4.50)]
G3 = [("G", "Belgio", "Egitto", 1.85), ("G", "Belgio", "Iran", 2.50), ("G", "Belgio", "Nuova Zelanda", 7.50)]
K2 = [("K", "Portogallo", "Colombia", 1.25), ("K", "Portogallo", "Rd Congo", 4.50)]

bollette = {
    1: {"stake": 0.10, "ncomb": 54, "imp": 5.40, "vmin": 82.14, "vmax": 5328.24, "sel":
        [("C", "Brasile", "Scozia", 2.25), ("E", "Germania", "Ecuador", 1.65)] + F3 + D3 +
        [("I", "Francia", "Senegal", 2.50), ("H", "Spagna", "Arabia Saudita", 3.50),
         ("G", "Belgio", "Egitto", 1.85), ("G", "Belgio", "Iran", 7.50), ("G", "Belgio", "Nuova Zelanda", 2.50),
         ("L", "Inghilterra", "Croazia", 1.35)] + K2 + [("J", "Argentina", "Austria", 1.60)]},
    2: {"stake": 0.10, "ncomb": 54, "imp": 5.40, "vmin": 79.65, "vmax": 5166.78, "sel":
        [("C", "Brasile", "Marocco", 1.60), ("E", "Germania", "Costa d'Avorio", 2.25)] + F3 + D3 +
        [("I", "Francia", "Senegal", 2.50), ("H", "Spagna", "Uruguay", 1.35)] + G3 +
        [("L", "Inghilterra", "Ghana", 3.50)] + K2 + [("J", "Argentina", "Austria", 1.60)]},
    3: {"stake": 0.15, "ncomb": 18, "imp": 2.70, "vmin": 369.47, "vmax": 5911.56, "sel":
        [("C", "Brasile", "Scozia", 2.25), ("E", "Germania", "Costa d'Avorio", 2.25)] + F3 + D3 +
        [("I", "Francia", "Senegal", 2.50), ("H", "Spagna", "Uruguay", 1.35),
         ("G", "Belgio", "Nuova Zelanda", 7.50), ("L", "Inghilterra", "Croazia", 1.35)] + K2 +
        [("J", "Argentina", "Algeria", 2.25)]},
    4: {"stake": 0.20, "ncomb": 36, "imp": 7.20, "vmin": 121.51, "vmax": 2627.36, "sel":
        [("C", "Brasile", "Scozia", 2.25), ("E", "Germania", "Costa d'Avorio", 2.25)] + F3 + D3 +
        [("I", "Francia", "Senegal", 2.50), ("H", "Spagna", "Uruguay", 1.35),
         ("G", "Belgio", "Iran", 2.50), ("G", "Belgio", "Egitto", 1.85),
         ("L", "Inghilterra", "Croazia", 1.35)] + K2 + [("J", "Argentina", "Algeria", 2.25)]},
    5: {"stake": 0.30, "ncomb": 54, "imp": 16.20, "vmin": 58.03, "vmax": 3822.63, "sel":
        [("C", "Brasile", "Marocco", 1.60), ("E", "Germania", "Ecuador", 1.65)] + F3 + D3 +
        [("I", "Francia", "Norvegia", 1.55), ("H", "Spagna", "Uruguay", 1.35)] + G3 +
        [("L", "Inghilterra", "Croazia", 1.35)] + K2 + [("J", "Argentina", "Algeria", 2.25)]},
    6: {"stake": 0.20, "ncomb": 54, "imp": 10.80, "vmin": 101.85, "vmax": 6607.02, "sel":
        [("C", "Brasile", "Scozia", 2.25), ("E", "Germania", "Ecuador", 1.65)] + F3 + D3 +
        [("I", "Francia", "Norvegia", 1.55), ("H", "Spagna", "Uruguay", 1.35)] + G3 +
        [("L", "Inghilterra", "Ghana", 3.50)] + K2 + [("J", "Argentina", "Austria", 1.60)]},
    7: {"stake": 0.20, "ncomb": 54, "imp": 10.80, "vmin": 98.77, "vmax": 6406.81, "sel":
        [("C", "Brasile", "Marocco", 1.60), ("E", "Germania", "Costa d'Avorio", 2.25)] + F3 + D3 +
        [("I", "Francia", "Norvegia", 1.55), ("H", "Spagna", "Arabia Saudita", 3.50)] + G3 +
        [("L", "Inghilterra", "Croazia", 1.35)] + K2 + [("J", "Argentina", "Austria", 1.60)]},
    8: {"stake": 0.05, "ncomb": 54, "imp": 2.70, "vmin": 128.59, "vmax": 8211.86, "sel":
        [("C", "Brasile", "Scozia", 2.25), ("E", "Germania", "Costa d'Avorio", 2.25)] + F3 + D3 +
        [("I", "Francia", "Norvegia", 1.55), ("H", "Spagna", "Arabia Saudita", 3.50)] + G3 +
        [("L", "Inghilterra", "Ghana", 3.50)] + K2 + [("J", "Argentina", "Algeria", 2.25)]},
    9: {"stake": 0.05, "ncomb": 54, "imp": 2.70, "vmin": 108.48, "vmax": 6906.99, "sel":
        [("C", "Brasile", "Marocco", 1.60), ("E", "Germania", "Ecuador", 1.65)] + F3 + D3 +
        [("I", "Francia", "Senegal", 2.50), ("H", "Spagna", "Arabia Saudita", 3.50)] + G3 +
        [("L", "Inghilterra", "Ghana", 3.50)] + K2 + [("J", "Argentina", "Algeria", 2.25)]},
}

# composizione ufficiale (SofaScore) - i nomi DEVONO combaciare con quelli delle bollette
GROUPS = [
    ("A", ["Messico", "Sudafrica", "Corea del Sud", "Rep. Ceca"]),
    ("B", ["Canada", "Bosnia", "Qatar", "Svizzera"]),
    ("C", ["Brasile", "Scozia", "Marocco", "Haiti"]),
    ("D", ["Stati Uniti", "Turchia", "Paraguay", "Australia"]),
    ("E", ["Germania", "Ecuador", "Costa d'Avorio", "Curaçao"]),
    ("F", ["Olanda", "Giappone", "Svezia", "Tunisia"]),
    ("G", ["Belgio", "Egitto", "Iran", "Nuova Zelanda"]),
    ("H", ["Spagna", "Uruguay", "Arabia Saudita", "Capo Verde"]),
    ("I", ["Francia", "Senegal", "Norvegia", "Iraq"]),
    ("J", ["Argentina", "Austria", "Algeria", "Giordania"]),
    ("K", ["Portogallo", "Colombia", "Rd Congo", "Uzbekistan"]),
    ("L", ["Inghilterra", "Croazia", "Ghana", "Panama"]),
]
GROUP_ORDER = [g for g, _ in GROUPS]
ALL_TEAMS = [(g, t) for g, ts in GROUPS for t in ts]

# righe dinamiche
T0 = 4
TN = T0 + len(ALL_TEAMS) - 1          # ultima riga squadre in Classifiche
G0, GN = 4, 4 + len(GROUP_ORDER) - 1  # tabella stato gironi (cols U/V)
A_ = "$A$%d:$A$%d" % (T0, TN)
H_ = "$H$%d:$H$%d" % (T0, TN)
I_ = "$I$%d:$I$%d" % (T0, TN)
K_ = "$K$%d:$K$%d" % (T0, TN)

# ---------------------------------------------------------------- STILI
HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE = Font(bold=True, size=14, color="1F4E78")
CENTER = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
GREEN_S = PatternFill("solid", fgColor="63BE7B")
GREEN_L = PatternFill("solid", fgColor="D9F2E1")   # verde tenue ma visibile
RED_S = PatternFill("solid", fgColor="E06666")
RED_L = PatternFill("solid", fgColor="F9D6D6")     # rosso tenue
PINK_L = PatternFill("solid", fgColor="FBD9E6")    # rosa tenue
YELLOW_L = PatternFill("solid", fgColor="FCF1B8")  # giallo tenue (un po' piu' carico, sennò invisibile)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

wb = openpyxl.Workbook()

# ================================================================ SCOMMESSE
ws = wb.active
ws.title = "Scommesse"
hdr = ["N. bolletta", "Indice", "Gruppo", "Prima squadra", "Seconda squadra", "Quota",
       "Pos 1a", "Pos 2a", "Qual 1a", "Qual 2a", "Elim 1a", "Elim 2a", "Girone chiuso",
       "Entrambe", "STATO", "Definitivo", "_q", "_dead", "_code"]
for c, h in enumerate(hdr, 1):
    cell = ws.cell(1, c, h)
    cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER

CB = "Classifiche!$B$%d:$B$%d" % (T0, TN)
CL_POS = "Classifiche!$L$%d:$L$%d" % (T0, TN)
CL_QUAL = "Classifiche!$M$%d:$M$%d" % (T0, TN)
CL_ELIM = "Classifiche!$O$%d:$O$%d" % (T0, TN)
CL_GU = "Classifiche!$U$%d:$U$%d" % (G0, GN)
CL_GV = "Classifiche!$V$%d:$V$%d" % (G0, GN)

r = 2
for b in sorted(bollette):
    idx = 0
    for (grp, p1, p2, q) in bollette[b]["sel"]:
        idx += 1
        ws.cell(r, 1, b); ws.cell(r, 2, idx); ws.cell(r, 3, grp)
        ws.cell(r, 4, p1); ws.cell(r, 5, p2)
        ws.cell(r, 6, q).number_format = "0.00"
        ws.cell(r, 7, '=IFERROR(INDEX(%s,MATCH(D%d,%s,0)),"")' % (CL_POS, r, CB))
        ws.cell(r, 8, '=IFERROR(INDEX(%s,MATCH(E%d,%s,0)),"")' % (CL_POS, r, CB))
        ws.cell(r, 9, '=IFERROR(INDEX(%s,MATCH(D%d,%s,0)),FALSE)' % (CL_QUAL, r, CB))
        ws.cell(r, 10, '=IFERROR(INDEX(%s,MATCH(E%d,%s,0)),FALSE)' % (CL_QUAL, r, CB))
        ws.cell(r, 11, '=IFERROR(INDEX(%s,MATCH(D%d,%s,0)),FALSE)' % (CL_ELIM, r, CB))
        ws.cell(r, 12, '=IFERROR(INDEX(%s,MATCH(E%d,%s,0)),FALSE)' % (CL_ELIM, r, CB))
        ws.cell(r, 13, '=IFERROR(INDEX(%s,MATCH(C%d,%s,0)),FALSE)' % (CL_GV, r, CL_GU))
        ws.cell(r, 14, '=AND(I%d,J%d)' % (r, r))
        ws.cell(r, 15, '=IF(N%d,IF(M%d,"VINTA (def.)","In vincita"),IF(OR(K%d,L%d,M%d),"PERSA (def.)","In perdita"))' % (r, r, r, r, r))
        ws.cell(r, 16, '=IF(OR(AND(N%d,M%d),AND(NOT(N%d),OR(K%d,L%d,M%d))),"SI","NO")' % (r, r, r, r, r, r))
        ws.cell(r, 17, '=IF(N%d,1,0)' % r)
        ws.cell(r, 18, '=IF(OR(K%d,L%d,AND(M%d,NOT(N%d))),1,0)' % (r, r, r, r))
        ws.cell(r, 19, '=IF(N%d,IF(M%d,2,1),IF(OR(K%d,L%d,M%d),-2,-1))' % (r, r, r, r, r))
        for c in range(1, 17):
            ws.cell(r, c).border = BORDER
            if c not in (4, 5, 15):
                ws.cell(r, c).alignment = CENTER
        r += 1
LAST = r - 1

rng = "A2:P%d" % LAST
ws.conditional_formatting.add(rng, FormulaRule(formula=['$S2=2'], fill=GREEN_S, stopIfTrue=True))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$S2=1'], fill=GREEN_L, stopIfTrue=True))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$S2=-2'], fill=RED_S, stopIfTrue=True))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$S2=-1'], fill=RED_L, stopIfTrue=True))
for i, w in enumerate([11, 7, 8, 16, 16, 7, 7, 7, 8, 8, 8, 8, 13, 9, 16, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for col in ("Q", "R", "S"):
    ws.column_dimensions[col].hidden = True
ws.freeze_panes = "A2"; ws.sheet_view.showGridLines = False

# ================================================================ CLASSIFICHE
cl = wb.create_sheet("Classifiche")
cl.cell(1, 1, "CLASSIFICHE GIRONI - Mondiali 2026").font = TITLE
cl.cell(1, 11, "Ultimo aggiornamento:").font = Font(bold=True)
lu = cl.cell(1, 13, "=Config!$B$2"); lu.number_format = "dd/mm/yyyy hh:mm:ss"; lu.font = Font(bold=True, color="1F4E78")
cl.cell(2, 11, "Stato:").font = Font(bold=True)
cl.cell(2, 13, "=Config!$B$3").font = Font(italic=True, color="1F4E78")
cl.cell(2, 1, "Dati live da football-data.org (Apps Script). In modalita' SIM (Config!B1=ON) i risultati NON vengono sovrascritti: modificali a mano nelle celle gialle.").font = Font(italic=True, color="808080")
clh = ["Gruppo", "Squadra", "V", "N", "P", "GF", "GS", "PG", "Punti", "DR", "Score",
       "Pos", "Qualif.", "MaxPti", "Elim.", "Esito", "Obbligo scommesse", "_app", "_tot"]
for c, h in enumerate(clh, 1):
    cell = cl.cell(3, c, h)
    cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER

row = T0
for grp, teams in GROUPS:
    for t in teams:
        cl.cell(row, 1, grp); cl.cell(row, 2, t)
        for c in (3, 4, 5, 6, 7):
            cell = cl.cell(row, c, 0); cell.fill = INPUT_FILL; cell.alignment = CENTER
        cl.cell(row, 8, '=C%d+D%d+E%d' % (row, row, row))
        cl.cell(row, 9, '=C%d*3+D%d' % (row, row))
        cl.cell(row, 10, '=F%d-G%d' % (row, row))
        # Score con tie-break deterministico (ordine in tabella) -> posizioni SEMPRE distinte 1-4,
        # quindi esattamente 2 "prime" anche a parita' totale.
        cl.cell(row, 11, '=$I%d*1000000000+($J%d+1000)*100000+$F%d*100+(%d-ROW())' % (row, row, row, TN))
        # Pos: ufficiale dall'API se presente; altrimenti rango per Score (distinto 1-4)
        cl.cell(row, 12, '=IF($T%d>0,$T%d,SUMPRODUCT((%s=$A%d)*(%s>$K%d))+1)' % (row, row, A_, row, K_, row))
        cl.cell(row, 13, '=($L%d<=2)' % row)
        cl.cell(row, 14, '=$I%d+(3-$H%d)*3' % (row, row))
        cl.cell(row, 15, '=SUMPRODUCT((%s=$A%d)*(%s>$N%d))>=2' % (A_, row, I_, row))
        cl.cell(row, 16, ('=IF(SUMPRODUCT((%s=$A{r})*(%s<3))=0,'
                          'IF($M{r},"Qualificata","Eliminata"),'
                          'IF($O{r},"Fuori dai giochi",IF($M{r},"In zona qualif.","In corsa")))'
                          % (A_, H_)).format(r=row))
        cl.cell(row, 17, '=IF($S%d=0,"",IF($R%d=0,"NON deve passare",IF($R%d>=$S%d,"DEVE passare","contesa")))' % (row, row, row, row))
        cl.cell(row, 18, '=COUNTIF(Scommesse!$D$2:$D$%d,$B%d)+COUNTIF(Scommesse!$E$2:$E$%d,$B%d)' % (LAST, row, LAST, row))
        cl.cell(row, 19, '=COUNTIF(Scommesse!$C$2:$C$%d,$A%d)' % (LAST, row))
        for c in range(1, 18):
            cl.cell(row, c).border = BORDER
            if c not in (2, 16, 17):
                cl.cell(row, c).alignment = CENTER
        row += 1

cl.cell(3, 21, "Gruppo").font = HDR; cl.cell(3, 21).fill = HDR_FILL; cl.cell(3, 21).alignment = CENTER
cl.cell(3, 22, "Completato").font = HDR; cl.cell(3, 22).fill = HDR_FILL; cl.cell(3, 22).alignment = CENTER
for i, g in enumerate(GROUP_ORDER):
    rr = G0 + i
    cl.cell(rr, 21, g).alignment = CENTER
    cl.cell(rr, 22, '=SUMPRODUCT((%s=$U%d)*(%s<3))=0' % (A_, rr, H_)).alignment = CENTER
    cl.cell(rr, 21).border = BORDER; cl.cell(rr, 22).border = BORDER

cl.conditional_formatting.add("A%d:Q%d" % (T0, TN), FormulaRule(formula=['$M%d=TRUE' % T0], fill=GREEN_L, stopIfTrue=False))
cl.conditional_formatting.add("O%d:O%d" % (T0, TN), FormulaRule(formula=['$O%d=TRUE' % T0], fill=RED_L, stopIfTrue=False))
cl.conditional_formatting.add("Q%d:Q%d" % (T0, TN), FormulaRule(formula=['ISNUMBER(SEARCH("DEVE",$Q%d))' % T0], fill=GREEN_S, font=Font(bold=True), stopIfTrue=True))
cl.conditional_formatting.add("Q%d:Q%d" % (T0, TN), FormulaRule(formula=['ISNUMBER(SEARCH("NON",$Q%d))' % T0], fill=RED_S, font=Font(bold=True, color="FFFFFF"), stopIfTrue=True))
for i, w in enumerate([8, 16, 4, 4, 4, 5, 5, 5, 6, 5, 7, 5, 8, 7, 7, 16, 18, 6, 6], 1):
    cl.column_dimensions[get_column_letter(i)].width = w
cl.column_dimensions["R"].hidden = True; cl.column_dimensions["S"].hidden = True
cl.column_dimensions["U"].width = 8; cl.column_dimensions["V"].width = 11

# colonne specchio (nascoste) per la vista "Gironi": Pos, Squadra, Pti, Rim(da giocare), icona
for row in range(T0, TN + 1):
    cl.cell(row, 24, '=$L%d' % row)            # Pos (effettiva)
    cl.cell(row, 25, '=$B%d' % row)            # Squadra
    cl.cell(row, 26, '=$I%d' % row)            # Pti
    cl.cell(row, 27, '=MAX(0,3-$H%d)' % row)   # partite ancora da giocare (3 - giocate)
    cl.cell(row, 28, '=IF($S%d=0,"",IF($R%d=0,"⛔",IF($R%d>=$S%d,"✅","·")))' % (row, row, row, row))  # icona
for col in range(24, 29):
    cl.column_dimensions[get_column_letter(col)].hidden = True
# posizione ufficiale scritta dall'API (col T); se assente, Pos usa il calcolo Punti->DR->GF
cl.cell(3, 20, "PosUff").font = HDR; cl.cell(3, 20).fill = HDR_FILL
cl.column_dimensions["T"].hidden = True
# codice obbligo (col AC) per la CF del Gironi: +1 deve passare, -1 non deve passare, 0 altro
for row in range(T0, TN + 1):
    cl.cell(row, 29, '=IF($S%d=0,0,IF($R%d=0,-1,IF($R%d>=$S%d,1,0)))' % (row, row, row, row))
cl.column_dimensions["AC"].hidden = True

cl.freeze_panes = "A%d" % T0; cl.sheet_view.showGridLines = False

# ================================================================ CALCOLI (nascosto)
ca = wb.create_sheet("Calcoli")
for c, h in enumerate(["Bolletta", "Gruppo", "nSel", "satNow", "closed", "satClosed", "deadSel", "unsat", "code"], 1):
    ca.cell(1, c, h).font = Font(bold=True)
rr = 2
calcoli_row = {}
for b in sorted(bollette):
    for g in GROUP_ORDER:
        if g not in {s[0] for s in bollette[b]["sel"]}:
            continue
        calcoli_row[(b, g)] = rr
        ca.cell(rr, 1, b); ca.cell(rr, 2, g)
        ca.cell(rr, 3, '=COUNTIFS(Scommesse!$A$2:$A$%d,$A%d,Scommesse!$C$2:$C$%d,$B%d)' % (LAST, rr, LAST, rr))
        ca.cell(rr, 4, '=IF(SUMIFS(Scommesse!$Q$2:$Q$%d,Scommesse!$A$2:$A$%d,$A%d,Scommesse!$C$2:$C$%d,$B%d)>0,1,0)' % (LAST, LAST, rr, LAST, rr))
        ca.cell(rr, 5, '=IFERROR(INDEX(%s,MATCH($B%d,%s,0)),FALSE)' % (CL_GV, rr, CL_GU))
        ca.cell(rr, 6, '=IF(AND($D%d=1,$E%d),1,0)' % (rr, rr))
        ca.cell(rr, 7, '=SUMIFS(Scommesse!$R$2:$R$%d,Scommesse!$A$2:$A$%d,$A%d,Scommesse!$C$2:$C$%d,$B%d)' % (LAST, LAST, rr, LAST, rr))
        ca.cell(rr, 8, '=IF(AND($C%d>0,$G%d=$C%d),1,0)' % (rr, rr, rr))
        # codice esito girone: 2=vinto def, -2=perso def, 1=in vincita, -1=in perdita
        ca.cell(rr, 9, '=IF($F%d=1,2,IF($H%d=1,-2,IF($D%d=1,1,-1)))' % (rr, rr, rr))
        rr += 1
CA_LAST = rr - 1
ca.sheet_state = "hidden"

# ================================================================ RIEPILOGO
rp = wb.create_sheet("Riepilogo")
rp.cell(1, 1, "RIEPILOGO BOLLETTE - stato live").font = TITLE
rp.cell(1, 6, "Ultimo aggiornamento:").font = Font(bold=True)
lur = rp.cell(1, 8, "=Config!$B$2"); lur.number_format = "dd/mm/yyyy hh:mm:ss"; lur.font = Font(bold=True, color="1F4E78")
rp.cell(2, 6, "Stato:").font = Font(bold=True)
rp.cell(2, 8, "=Config!$B$3").font = Font(italic=True, color="1F4E78")
for c, h in enumerate(["Bolletta", "Punt./comb", "N. comb.", "Importo", "Vinc. min", "Vinc. max",
                       "Gironi OK", "STATO", "Definitivo", "Vincita attuale stimata",
                       "_tot", "_satNow", "_satClosed", "_unsat", "_code"], 1):
    cell = rp.cell(3, c, h)
    cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CENTER; cell.border = BORDER

row = 4
for b in sorted(bollette):
    m = bollette[b]
    rp.cell(row, 1, "Bolletta %d" % b)
    rp.cell(row, 2, m["stake"]).number_format = "0.00"
    rp.cell(row, 3, m["ncomb"])
    rp.cell(row, 4, m["imp"]).number_format = '#,##0.00 "EUR"'
    rp.cell(row, 5, m["vmin"]).number_format = '#,##0.00 "EUR"'
    rp.cell(row, 6, m["vmax"]).number_format = '#,##0.00 "EUR"'
    bexpr = "VALUE(MID($A%d,10,3))" % row
    rp.cell(row, 11, '=COUNTIFS(Calcoli!$A$2:$A$%d,%s)' % (CA_LAST, bexpr))
    rp.cell(row, 12, '=SUMIFS(Calcoli!$D$2:$D$%d,Calcoli!$A$2:$A$%d,%s)' % (CA_LAST, CA_LAST, bexpr))
    rp.cell(row, 13, '=SUMIFS(Calcoli!$F$2:$F$%d,Calcoli!$A$2:$A$%d,%s)' % (CA_LAST, CA_LAST, bexpr))
    rp.cell(row, 14, '=SUMIFS(Calcoli!$H$2:$H$%d,Calcoli!$A$2:$A$%d,%s)' % (CA_LAST, CA_LAST, bexpr))
    tot, sat, satc, uns = "$K%d" % row, "$L%d" % row, "$M%d" % row, "$N%d" % row
    rp.cell(row, 7, '=%s&"/"&%s' % (sat, tot))
    rp.cell(row, 8, '=IF(%s=%s,"VINCENTE (def.)",IF(%s>0,"PERDENTE (def.)",IF(%s=%s,"Vincente (provv.)","Perdente (provv.)")))' % (satc, tot, uns, sat, tot))
    rp.cell(row, 9, '=IF(OR(%s=%s,%s>0),"SI","NO")' % (satc, tot, uns))
    rp.cell(row, 10, ('=IFERROR(IF(%s=%s,$B%d*EXP(SUMPRODUCT((Scommesse!$A$2:$A$%d=%s)*'
                      '(Scommesse!$Q$2:$Q$%d)*LN(Scommesse!$F$2:$F$%d))),"-"),"-")'
                      % (sat, tot, row, LAST, bexpr, LAST, LAST))).number_format = '#,##0.00 "EUR"'
    rp.cell(row, 15, '=IF(%s=%s,2,IF(%s>0,-2,IF(%s=%s,1,-1)))' % (satc, tot, uns, sat, tot))
    for c in range(1, 11):
        rp.cell(row, c).border = BORDER
        if c not in (1, 8):
            rp.cell(row, c).alignment = CENTER
    row += 1
RP_LAST = row - 1
rp.conditional_formatting.add("A4:J%d" % RP_LAST, FormulaRule(formula=['$O4=2'], fill=GREEN_S, stopIfTrue=True))
rp.conditional_formatting.add("A4:J%d" % RP_LAST, FormulaRule(formula=['$O4=1'], fill=GREEN_L, stopIfTrue=True))
rp.conditional_formatting.add("A4:J%d" % RP_LAST, FormulaRule(formula=['$O4=-2'], fill=RED_S, stopIfTrue=True))
rp.conditional_formatting.add("A4:J%d" % RP_LAST, FormulaRule(formula=['$O4=-1'], fill=RED_L, stopIfTrue=True))
for i, w in enumerate([11, 11, 9, 12, 12, 12, 10, 18, 10, 22], 1):
    rp.column_dimensions[get_column_letter(i)].width = w
for col in ("K", "L", "M", "N", "O"):
    rp.column_dimensions[col].hidden = True
rp.sheet_view.showGridLines = False

# ================================================================ CONFIG
cf = wb.create_sheet("Config")
cf.column_dimensions["A"].width = 30; cf.column_dimensions["B"].width = 28
cf.cell(1, 1, "Modalita' simulazione (SIM)").font = Font(bold=True)
b1 = cf.cell(1, 2, "OFF"); b1.fill = INPUT_FILL; b1.alignment = CENTER; b1.font = Font(bold=True)
dv = DataValidation(type="list", formula1='"OFF,ON"', allow_blank=False)
cf.add_data_validation(dv); dv.add(b1)
cf.cell(2, 1, "Ultimo aggiornamento dati").font = Font(bold=True)
cf.cell(2, 2, "(in attesa dello script)")
cf.cell(3, 1, "Stato aggiornamento").font = Font(bold=True)
cf.cell(3, 2, "(in attesa dello script)")
cf.cell(4, 1, "Fonte dati").font = Font(bold=True); cf.cell(4, 2, "football-data.org - World Cup")
note = ["", "SIM = ON  -> lo script NON sovrascrive: puoi simulare i risultati a mano nelle celle gialle del foglio Classifiche.",
        "SIM = OFF -> lo script aggiorna i risultati reali da football-data.org, ma SOLO durante le partite",
        "            (finestra: inizio partita -> +155 min). Fuori dalle partite resta fermo (vedi 'Stato aggiornamento').",
        "Per attivare il live: vedi foglio Istruzioni (Estensioni > Apps Script)."]
for i, t in enumerate(note, 6):
    cf.cell(i, 1, t).font = Font(italic=True, color="808080")
cf.sheet_view.showGridLines = False

# ================================================================ ISTRUZIONI
iz = wb.create_sheet("Istruzioni")
iz.column_dimensions["A"].width = 125
lines = [
    ("SCOMMESSE MONDIALI 2026 - GUIDA", TITLE),
    ("", None),
    ("1) CARICA SU GOOGLE DRIVE: trascina questo file su drive.google.com, poi aprilo e 'Apri con Fogli Google'", None),
    ("   (oppure File > Salva come Fogli Google). Le formule funzionano subito.", None),
    ("", None),
    ("2) DATI LIVE (auto-aggiornamento) - fonte: football-data.org:", Font(bold=True, size=12, color="1F4E78")),
    ("   a. Registrati gratis su football-data.org/client/register e ottieni un token via email.", None),
    ("   b. Menu  Estensioni > Apps Script ; incolla il file Codice.gs e metti il token in API_TOKEN. Salva.", None),
    ("   c. Esegui una volta 'updateStandings' e autorizza i permessi richiesti.", None),
    ("   d. Riapri il foglio: menu '⚽ Scommesse' > 'Installa aggiornamento automatico'.", None),
    ("   -> I risultati si aggiornano da soli ogni 15 min MA SOLO durante le partite (vedi 'Stato' qui sopra).", None),
    ("      I collaboratori col link vedono i dati in tempo reale, senza consumare chiamate API.", None),
    ("", None),
    ("3) CONDIVIDI: pulsante Condividi (in alto a destra) > 'Chiunque abbia il link' > Visualizzatore. Copia il link.", None),
    ("", None),
    ("4) SIMULAZIONE: nel foglio Config metti  SIM = ON. Lo script smette di sovrascrivere; nel foglio Classifiche", None),
    ("   modifica V/N/P/GF/GS (celle gialle) e guarda Scommesse/Riepilogo cambiare in tempo reale (vinci/perdi).", None),
    ("   Per tornare ai dati reali: Config SIM = OFF (lo script riprende durante le partite).", None),
    ("", None),
    ("LETTURA COLORI: verde=in vincita/vinta, rosso=in perdita/persa, 'def.'=esito non piu' modificabile.", None),
    ("Classifiche: ✅ DEVE passare = squadra in tutte le tue coppie del girone; ⛔ NON deve passare = mai giocata.", None),
    ("Una bolletta vince solo se TUTTI e 10 i gironi hanno la coppia corretta (primi 2 del girone).", None),
    ("Nota: quota Belgio/Nuova Zelanda della Bolletta 1 stimata a 2,50 (era illeggibile sullo scontrino).", None),
]
for i, (txt, fnt) in enumerate(lines, 1):
    cell = iz.cell(i, 1, txt)
    if fnt:
        cell.font = fnt
iz.sheet_view.showGridLines = False

# ================================================================ GIRONI (vista)
gv = wb.create_sheet("Gironi", 0)          # prima scheda: landing di sola lettura
gv.sheet_view.showGridLines = False
gv.cell(1, 1, "GIRONI - Classifiche live").font = TITLE
gv.sheet_view.zoomScale = 70
gv.cell(1, 14, "Aggiornato:").font = Font(bold=True)
g_lu = gv.cell(1, 16, "=Config!$B$2"); g_lu.number_format = "dd/mm/yyyy hh:mm"; g_lu.font = Font(bold=True, color="1F4E78")
gv.cell(2, 1, "Verde = situazione OK per le scommesse (✅ tra i primi 2 / ⛔ fuori dai primi 2).  Rosa = situazione sfavorevole.  Rim = partite da giocare.").font = Font(italic=True, color="808080")

LEFTAL = Alignment(horizontal="left", vertical="center")
HEAD = ["Pos", "Squadra", "Pti", "Rim", ""]
NCOLS = 5
CARDW = NCOLS + 1     # 5 dati + 1 gap
CARDH = 6 + 1         # titolo + header + 4 squadre + 1 gap
TOP = 4
for k, (grp, _t) in enumerate(GROUPS):
    cr, cc = k // 6, k % 6          # 6 colonne x 2 righe (16:9)
    r0 = TOP + cr * CARDH
    c0 = 1 + cc * CARDW
    gv.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + NCOLS - 1)
    tc = gv.cell(r0, c0, "GIRONE %s" % grp)
    tc.font = Font(bold=True, color="FFFFFF"); tc.fill = HDR_FILL; tc.alignment = CENTER
    for j, h in enumerate(HEAD):
        hc = gv.cell(r0 + 1, c0 + j, h); hc.font = Font(bold=True, size=9, color="595959"); hc.alignment = CENTER
    # una formula DIRETTA per cella (niente spill): la squadra in posizione 'rank' del girone.
    # Cosi' la formattazione condizionale si applica davvero (lo spill non la attiva).
    for rank in range(1, 5):
        rr = r0 + 1 + rank
        match = ('MATCH(1,(Classifiche!$A$%d:$A$%d="%s")*(Classifiche!$L$%d:$L$%d=%d),0)'
                 % (T0, TN, grp, T0, TN, rank))
        gv.cell(rr, c0, rank)                                                              # Pos (1..4)
        gv.cell(rr, c0 + 1, '=IFERROR(INDEX(Classifiche!$B$%d:$B$%d,%s),"")' % (T0, TN, match))   # Squadra
        gv.cell(rr, c0 + 2, '=IFERROR(INDEX(Classifiche!$I$%d:$I$%d,%s),"")' % (T0, TN, match))    # Pti
        gv.cell(rr, c0 + 3, '=IFERROR(INDEX(Classifiche!$AA$%d:$AA$%d,%s),"")' % (T0, TN, match))  # Rim
        gv.cell(rr, c0 + 4, '=IFERROR(INDEX(Classifiche!$AB$%d:$AB$%d,%s),"")' % (T0, TN, match))  # icona
        gv.cell(rr, 40 + cc, '=IFERROR(INDEX(Classifiche!$AC$%d:$AC$%d,%s),0)' % (T0, TN, match))  # codice obbligo (helper)
        for j in range(NCOLS):
            cell = gv.cell(rr, c0 + j); cell.border = BORDER
            cell.alignment = LEFTAL if j == 1 else CENTER
    posL = get_column_letter(c0)
    iconL = get_column_letter(c0 + 4)
    rng = "%s%d:%s%d" % (posL, r0 + 2, iconL, r0 + 5)
    # CF puramente numerica (come nel Tabellone): legge il codice obbligo dalla colonna helper
    # nascosta (+1 deve passare, -1 non deve). Niente emoji ne' funzioni recenti nella regola.
    obb = '$%s%d' % (get_column_letter(40 + cc), r0 + 2)
    pos = '$%s%d' % (posL, r0 + 2)
    # verde = obbligo rispettato
    gv.conditional_formatting.add(rng, FormulaRule(
        formula=['=OR(AND(%s=1,%s<=2),AND(%s=-1,%s>2))' % (obb, pos, obb, pos)], fill=GREEN_L, stopIfTrue=True))
    # rosa = situazione opposta
    gv.conditional_formatting.add(rng, FormulaRule(
        formula=['=OR(AND(%s=1,%s>2),AND(%s=-1,%s<=2))' % (obb, pos, obb, pos)], fill=PINK_L, stopIfTrue=True))
for cc in range(6):
    c0 = 1 + cc * CARDW
    gv.column_dimensions[get_column_letter(c0)].width = 4
    gv.column_dimensions[get_column_letter(c0 + 1)].width = 14
    gv.column_dimensions[get_column_letter(c0 + 2)].width = 5
    gv.column_dimensions[get_column_letter(c0 + 3)].width = 4
    gv.column_dimensions[get_column_letter(c0 + 4)].width = 4
    gv.column_dimensions[get_column_letter(c0 + 5)].width = 2
for cc in range(6):
    gv.column_dimensions[get_column_letter(40 + cc)].hidden = True   # colonne helper codice obbligo

# ================================================================ TABELLONE (vista pubblica bollette)
# righe Scommesse per bolletta (stesso ordine con cui sono state scritte)
bol_rows = {}
_rr = 2
for b in sorted(bollette):
    bol_rows[b] = []
    for _ in bollette[b]["sel"]:
        bol_rows[b].append(_rr); _rr += 1

tb = wb.create_sheet("Tabellone", 0)          # vista pubblica, in primo piano
tb.sheet_view.showGridLines = False
tb.sheet_view.zoomScale = 70
tb.cell(1, 1, "TABELLONE BOLLETTE - live").font = TITLE
tb.cell(2, 1, "Per ogni scommessa: ✔ entrambe passano / ✖ no.  Sfondo pieno = esito DEFINITIVO (verde vinta, rosso persa).").font = Font(italic=True, color="808080")
tb.cell(1, 28, "Aggiornato:").font = Font(bold=True)
t_lu = tb.cell(1, 30, "=Config!$B$2"); t_lu.number_format = "dd/mm/yyyy hh:mm"; t_lu.font = Font(bold=True, color="1F4E78")

MAXSEL = 17
TOP_T = 4
CARDCOLS_T = 6        # Gir, Coppia, Quota, Icona, code(nascosto), gap
SUBH = ["Gir", "Coppia", "Q", "✔"]
for bi, b in enumerate(sorted(bollette)):
    c0 = 1 + bi * CARDCOLS_T
    codecol = c0 + 4
    cL = get_column_letter(codecol)
    sels = bollette[b]["sel"]
    # titolo con importo
    tb.merge_cells(start_row=TOP_T, start_column=c0, end_row=TOP_T, end_column=c0 + 3)
    tt = tb.cell(TOP_T, c0, "BOLLETTA %d  (%.2f€)" % (b, bollette[b]["imp"]))
    tt.font = Font(bold=True, color="FFFFFF"); tt.fill = HDR_FILL; tt.alignment = CENTER
    for j, h in enumerate(SUBH):
        hc = tb.cell(TOP_T + 1, c0 + j, h); hc.font = Font(bold=True, size=9, color="595959"); hc.alignment = CENTER
    # esito di GIRONE per lo sfondo (codice da Calcoli); ✔ solo sulla coppia che vince davvero
    for i in range(MAXSEL):
        r = TOP_T + 2 + i
        if i < len(sels):
            srow = bol_rows[b][i]
            grp = sels[i][0]
            crow = calcoli_row[(b, grp)]
            tb.cell(r, c0, '=Scommesse!$C$%d' % srow)
            tb.cell(r, c0 + 1, '=Scommesse!$D$%d&" / "&Scommesse!$E$%d' % (srow, srow))
            tb.cell(r, c0 + 2, '=Scommesse!$F$%d' % srow).number_format = "0.00"
            tb.cell(r, codecol, '=Calcoli!$I$%d' % crow)              # codice esito di GIRONE (per sfondo)
            tb.cell(r, c0 + 3, '=IF(Scommesse!$S$%d>0,"✔","")' % srow)  # ✔ solo se questa coppia sta vincendo
        for j in range(4):
            cell = tb.cell(r, c0 + j); cell.border = BORDER; cell.font = Font(size=9)
            cell.alignment = LEFTAL if j == 1 else CENTER
    # contorno attorno alle righe dello stesso girone (raggruppamento visivo)
    med = Side(style="medium", color="9AA4B2")
    gi = 0
    while gi < len(sels):
        g0 = sels[gi][0]
        gj = gi
        while gj < len(sels) and sels[gj][0] == g0:
            gj += 1
        r1 = TOP_T + 2 + gi
        r2 = TOP_T + 2 + (gj - 1)
        for rr in range(r1, r2 + 1):
            for cc in range(c0, c0 + 4):
                tb.cell(rr, cc).border = Border(
                    left=med if cc == c0 else thin,
                    right=med if cc == c0 + 3 else thin,
                    top=med if rr == r1 else thin,
                    bottom=med if rr == r2 else thin)
        gi = gj
    # sfondo di GIRONE: verde=definitivo vinto, rosa=definitivo perso, giallo=aperto senza vincenti
    block = "%s%d:%s%d" % (get_column_letter(c0), TOP_T + 2, get_column_letter(c0 + 3), TOP_T + 1 + MAXSEL)
    tb.conditional_formatting.add(block, FormulaRule(formula=['$%s%d=2' % (cL, TOP_T + 2)], fill=GREEN_L, stopIfTrue=True))
    tb.conditional_formatting.add(block, FormulaRule(formula=['$%s%d=-2' % (cL, TOP_T + 2)], fill=PINK_L, stopIfTrue=True))
    tb.conditional_formatting.add(block, FormulaRule(formula=['$%s%d=-1' % (cL, TOP_T + 2)], fill=YELLOW_L, stopIfTrue=True))
    # footer: solo ✔/✖ + vincita ; sfondo verde/rosso SOLO se la bolletta e' definitiva
    rpRow = 3 + b
    footRow = TOP_T + 2 + MAXSEL
    tb.merge_cells(start_row=footRow, start_column=c0, end_row=footRow, end_column=c0 + 3)
    # vincita ATTUALE = puntata x prodotto delle sole quote vincenti (ignora le perdenti)
    stake_lit = '%.2f' % bollette[b]["stake"]
    vc = tb.cell(footRow, c0, ('=IF(Riepilogo!$O$%d>0,"✔ ","✖ ")&"Attuale "&TEXT(%s*EXP(SUMPRODUCT('
                               '(Scommesse!$A$2:$A$%d=%d)*(Scommesse!$Q$2:$Q$%d)*LN(Scommesse!$F$2:$F$%d))),"#,##0.00")&" €"')
                 % (rpRow, stake_lit, LAST, b, LAST, LAST))
    vc.alignment = CENTER; vc.font = Font(bold=True, color="1F4E78")
    tb.cell(footRow, codecol, '=Riepilogo!$O$%d' % rpRow)          # codice bolletta (nascosto, per sfondo)
    for j in range(4):
        tb.cell(footRow, c0 + j).border = BORDER
    frng = "%s%d:%s%d" % (get_column_letter(c0), footRow, get_column_letter(c0 + 3), footRow)
    tb.conditional_formatting.add(frng, FormulaRule(formula=['$%s%d=2' % (cL, footRow)], fill=GREEN_L, stopIfTrue=True))
    tb.conditional_formatting.add(frng, FormulaRule(formula=['$%s%d=-2' % (cL, footRow)], fill=RED_L, stopIfTrue=True))
    tb.column_dimensions[get_column_letter(c0)].width = 3.5
    tb.column_dimensions[get_column_letter(c0 + 1)].width = 18
    tb.column_dimensions[get_column_letter(c0 + 2)].width = 5
    tb.column_dimensions[get_column_letter(c0 + 3)].width = 3.5
    tb.column_dimensions[cL].hidden = True
    tb.column_dimensions[get_column_letter(c0 + 5)].width = 2

# i visualizzatori (sola lettura) vedono SOLO "Tabellone" e "Gironi"; le altre schede sono nascoste
# (restano accessibili all'editor e continuano a calcolare/aggiornarsi normalmente)
for name in ("Scommesse", "Classifiche", "Calcoli", "Riepilogo", "Config", "Istruzioni"):
    wb[name].sheet_state = "hidden"
wb.active = wb.sheetnames.index("Tabellone")     # apri sul Tabellone

_OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "dist")
os.makedirs(_OUT_DIR, exist_ok=True)
wb.save(os.path.join(_OUT_DIR, "Mondiali2026_Scommesse.xlsx"))
print("OK squadre:", len(ALL_TEAMS), "| righe scommesse:", LAST - 1, "| righe calcoli:", CA_LAST - 1)
